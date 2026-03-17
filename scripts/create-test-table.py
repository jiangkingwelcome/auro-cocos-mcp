import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 读取测试用例数据
with open('test-cases.json', 'r', encoding='utf-8') as f:
    test_cases = json.load(f)

# 创建工作簿
wb = Workbook()
sheet = wb.active
sheet.title = '测试用例'

# 定义表头
headers = ['测试ID', '工具名称', 'Action名称', '测试标题', '输入参数', '期望结果', '备注']
sheet.append(headers)

# 格式化表头
header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
header_font = Font(bold=True, size=11)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col in range(1, len(headers) + 1):
    cell = sheet.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# 填充数据
for case in test_cases:
    # 格式化输入参数为易读的 JSON
    input_json = json.dumps(case['input'], ensure_ascii=False, indent=2)

    row_data = [
        case['id'],
        case['tool'],
        case['action'],
        case['title'],
        input_json,
        case['expected'],
        case['note']
    ]
    sheet.append(row_data)

# 设置列宽和格式
column_widths = {
    'A': 10,  # 测试ID
    'B': 20,  # 工具名称
    'C': 25,  # Action名称
    'D': 30,  # 测试标题
    'E': 50,  # 输入参数
    'F': 50,  # 期望结果
    'G': 30   # 备注
}

for col_letter, width in column_widths.items():
    sheet.column_dimensions[col_letter].width = width

# 格式化数据行
data_alignment = Alignment(vertical='top', wrap_text=True)
for row in range(2, len(test_cases) + 2):
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=row, column=col)
        cell.alignment = data_alignment
        cell.border = thin_border

        # 输入参数列使用等宽字体
        if col == 5:
            cell.font = Font(name='Consolas', size=9)

# 添加筛选功能
sheet.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(test_cases) + 1}'

# 冻结首行
sheet.freeze_panes = 'A2'

# 保存文件
wb.save('test-cases.xlsx')
print(f'[OK] 已创建测试用例表格: test-cases.xlsx')
print(f'[OK] 共 {len(test_cases)} 个测试用例')
